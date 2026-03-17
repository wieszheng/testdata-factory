"""Excel 导出器"""

from typing import Any
import io


def export_to_excel(data: list[dict], columns: list[str], sheet_name: str = "Sheet1") -> bytes:
    """
    将数据导出为 Excel 格式
    
    Args:
        data: 数据列表
        columns: 列名列表
        sheet_name: 工作表名称
        
    Returns:
        Excel 文件的字节数据
    """
    try:
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 写入表头
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col)
        
        # 写入数据
        for row_idx, row in enumerate(data, start=2):
            for col_idx, col in enumerate(columns, start=1):
                value = row.get(col, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 写入到字节流
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")


def export_single_column_to_excel(data: list[str], column_name: str, sheet_name: str = "Sheet1") -> bytes:
    """
    将单列数据导出为 Excel 格式
    
    Args:
        data: 数据列表
        column_name: 列名
        sheet_name: 工作表名称
        
    Returns:
        Excel 文件的字节数据
    """
    return export_to_excel(
        [{column_name: v} for v in data],
        [column_name],
        sheet_name
    )
